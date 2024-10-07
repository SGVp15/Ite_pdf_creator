import copy
import os
import re

from openpyxl import load_workbook

from UTILS.log import log
from contact import Contact
from config import PAGE_NAME, map_excel_user, OUT_PATH, FILE_XLSX
from course import Course


def read_excel_file(filename=FILE_XLSX, sheet_names=('2015',)) -> {tuple}:
    workbook = load_workbook(filename=filename, read_only=True, data_only=True)
    all_data = {}
    for sheet_name in sheet_names:
        sheet = workbook[sheet_name]
        data = []
        for row in sheet.iter_rows(values_only=True):
            # data.append(list(map(clean_str, row)))
            data.append(row)
        all_data[sheet_name] = data
    workbook.close()
    return all_data


def read_excel(excel, column, row) -> str:
    sheet_ranges = excel[PAGE_NAME]
    return clean_str(str(sheet_ranges[f'{column}{row}'].value))


def clean_str(s):
    if type(s) is str:
        s = s.strip()
        s = s.replace(',', ', ')
        s = re.sub(r'\s{2,}', ' ', s)
        if s in ('None', '#N/A'):
            s = ''
    return s


def get_contact_from_excel(rows_excel, templates_docx) -> [Contact]:
    filename = FILE_XLSX
    file_excel = load_workbook(filename=filename, read_only=True, data_only=True)

    contacts_excel = []
    for i in rows_excel:
        from_excel = {}
        for k, v in map_excel_user.items():
            from_excel[k] = read_excel(file_excel, column=v, row=i)
        try:
            contacts_excel.append(Contact(from_excel))
        except TypeError:
            pass
    file_excel.close()

    contacts = []
    for contact_excel in contacts_excel:
        for template in templates_docx:
            contact: Contact = copy.deepcopy(contact_excel)
            contact.docx_list_files_name_templates = template
            for _path in ('pdf', 'docx'):
                path_folder = os.path.join(OUT_PATH, _path, contact.dir_name)
                os.makedirs(path_folder, exist_ok=True)

            contacts.append(contact)
    return contacts


def read_users_from_excel(file_excel=FILE_XLSX, header=False, rows_users=(-1,),
                          sheet_names=('2015', 'Курсы', 'Архив Курсов', 'Шаблоны')) -> [Contact]:
    data_excel = read_excel_file(file_excel, sheet_names=sheet_names)
    users_data = data_excel.get('2015')
    if rows_users != (-1,):
        users_data = [users_data[i - 1] for i in rows_users]
    elif header is False:
        users_data = users_data[1:]

    users_data = [u for u in users_data if u[1] is not None]

    courses_data: list = []
    courses_data.extend(data_excel.get('Курсы')[1:])
    courses_data.extend(data_excel.get('Архив Курсов')[1:])

    templates_data = data_excel.get('Шаблоны')[1:]

    courses = []
    for c in courses_data:
        try:
            courses.append(Course(c))
        except ValueError:
            pass

    templates = []
    for t in templates_data:
        templates.append(t[1])

    users = []
    for data in users_data:
        try:
            users.append(Contact(data, courses, templates))
        except ValueError:
            log.error(f'[DataError] {data}')
    users = [u for u in users if u.abr_course is not None and u.course is not None]
    return users
